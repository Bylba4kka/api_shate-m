import openpyxl
import requests
import json
import base64
import paramiko

from io import BytesIO
from PIL import Image
from openpyxl.drawing.image import Image as OpenpyxlImage

from config import API_KEY, HEIGHTSIZE, WIDTHSIZE, server_hostname, server_username, server_password

API_BASE_URL = "https://api.shate-m.ru/api/v1/articles"
AUTH_URL_LOGIN = "https://api.shate-m.ru/api/v1/auth/login"  # URL для получения токена по логину и паролю
AUTH_URL_APIKEY = "https://api.shate-m.ru/api/v1/auth/loginbyapikey"  # URL для получения токена по API ключу




def upload_via_ssh(local_path, remote_path, hostname, username, password):
    try:
        # Устанавливаем SSH-соединение
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_client.connect(hostname=hostname, username=username, password=password)

        # Создаем SFTP-соединение
        sftp_client = ssh_client.open_sftp()

        # Загружаем файл
        sftp_client.put(local_path, remote_path)

        # Закрываем соединение
        sftp_client.close()
        ssh_client.close()
        print(f"Файл {local_path} успешно загружен на {hostname} по пути {remote_path}")
    except Exception as e:
        print(f"Ошибка при загрузке файла: {str(e)}")


def get_access_token_by_apikey():
    auth_data = {
        'apikey': API_KEY,
    }
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    
    try:
        response = requests.post(AUTH_URL_APIKEY, data=auth_data, headers=headers)
        response.raise_for_status()
        data = response.json()
        return data.get('access_token')
    except requests.RequestException as e:
        print(f"Failed to obtain access token by API key: {e}")
        return None


def get_image_url(contentId, headers):
    payload = {"contentKeys": [
        {
            "contentId": contentId,
            "heightSize": HEIGHTSIZE,  
            "widthSize": WIDTHSIZE
        }
    ]}

    url = "https://api.shate-m.ru/api/v1/contents/search"
    response = requests.post(url, headers=headers, json=payload)
    data = json.loads(response.text)
    base64_image = data[0].get("value")

    return base64_image


def search_article(article_search_string, headers):
    search_url = f"{API_BASE_URL}/search/{article_search_string}"
    try:
        response = requests.get(search_url, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        if data and 'id' in data[0]["article"]:  # предполагаем, что API возвращает список с первым элементом
            return data[0]["article"]['id']
        else:
            print(f"No ID found for article {article_search_string}")
            return None
    except requests.RequestException as e:
        print(f"API request error: {e}")
        return None


def get_analogs(article_id, headers):
    analogs_url = f"{API_BASE_URL}/{article_id}/analogs"
    response = requests.get(analogs_url, headers=headers)
    response_data = json.loads(response.text)
    analogs = [item['article']['code'] for item in response_data if 'article' in item]
    return analogs


def decode_base64_image(base64_str):
    if "," in base64_str:
        base64_str = base64_str.split(",")[1]
    
    image_data = base64.b64decode(base64_str)
    
    image = Image.open(BytesIO(image_data))
    
    return image


def fetch_article_details(article_id, headers):
    article_url = f"{API_BASE_URL}/{article_id}?include=contents"
    try:
        response = requests.get(article_url, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        # Извлекаем URL изображения
        image_url = None
        if 'contents' in data:
            for content in data['contents']:
                if content.get('contentType') == 'ImageTwoDimensional':
                    image_url = get_image_url(content.get('contentId'), headers)
                    break
        
        return image_url
    

    except requests.RequestException as e:
        print(f"API request error: {e}")
        return None
    except KeyError:
        print("Unexpected API response format")
        return None

def process_excel(file_path, headers):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    counter = 0
    data_len = ws.max_row
    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):
        counter += 1
        article_number = row[0].value
        print(f"{counter} article ({article_number}) of {data_len}")

        if article_number:
            article_id = search_article(article_number, headers)

            if article_id:
                base64_image_str  = fetch_article_details(article_id, headers)

                if base64_image_str:
                    image = decode_base64_image(base64_image_str)
                    
                    
                if base64_image_str:
                    image = decode_base64_image(base64_image_str)
                    
                    local_file = "images/" + f"{article_number}_image.png"

                    remote_path = '/var/www/html/imgs/' + f"{article_number}_image.png"     

                    # Загрузка картинки на сервер
                    upload_via_ssh(local_file, remote_path, server_hostname, server_username, server_password)

                    # Создание openpyxl картинки
                    # img = OpenpyxlImage(image_path)

                    # # Добавление картинки в эксель таблицу
                    # ws.add_image(img, f"B{row[0].row}")
                    # Добавление ссылки на картинку в эксель таблицу

                    ws.cell(row=row[0].row, column=2, value="http://95.163.221.230/imgs/" + f"{article_number}_image.png")
                

                analogs = get_analogs(article_id, headers)
                if analogs:
                    ws.cell(row=row[0].row, column=3, value="\n".join(analogs))
    
    wb.save(file_path)
    print("Processing completed.")

if __name__ == "__main__":
    token = get_access_token_by_apikey()  # используйте get_access_token_by_login() если требуется
    # print(token)
    if token:
        headers = {
            'Authorization': f'Bearer {token}',
            "accept": "text/plain",
        }
        file_path = 'data.xlsx'
        process_excel(file_path, headers)
    else:
        print("Failed to obtain access token. Exiting...")


